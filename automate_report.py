from discord import Webhook
import pandas as pd #pandas untuk membuat dataframe
from openpyxl import load_workbook #openpyxl untuk interaksi antara python dan file excel
from openpyxl.styles import *
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList

input_file = 'data_set/supermarket_sales.xlsx'
output_file = 'data_set/report_2019.xlsx'
# output_file = 'data_set/report_2019.csv'
sheetname = 'Report Payment Method'
webhook_url = 'https://discord.com/api/webhooks/1026290023240306729/EFXrVrMHUHPZfTgTjqadcvELPmDly8n_h5DPmK4X3irZX1dh9w-cUMrwEOuKS7WoTKWX'

df = pd.read_excel(input_file)

#select columns dan filter by gender
# df = df[['Gender', 'Customer type', 'Unit price', 'Quantity', 'Total']].loc[df['Gender'] == 'Female'].head()

#pivoting selection columns
df_pivot = df.pivot_table(index='Gender', columns='Payment', values='Total', aggfunc='sum').round(decimals=2)
print(f'Dataframe columns: {df.columns}')
print('Sample dataset:')
df = df_pivot.apply(pd.to_numeric, errors='coerce')
print(f'{df}')
print()

#save result to excel
print('Save dataframe to excel.....')
df_pivot.to_excel(excel_writer=output_file, sheet_name=sheetname)
print('Save dataframe done.....')

#save result to csv
# print('Save dataframe to csv.....')
# df_pivot.to_csv(path_or_buf=output_file)
# print('Save dataframe done.....')

#grafik
wb = load_workbook(output_file)
wb.active = wb[sheetname]

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# print(min_column, max_column, min_row, max_row)

barchart = BarChart()

data = Reference(wb.active, 
                    min_col=min_column+1, 
                    max_col=max_column, 
                    min_row=min_row, 
                    max_row=max_row
                )

categories = Reference(wb.active, 
                        min_col=min_column, 
                        max_col=max_column, 
                        min_row=min_row+1, 
                        max_row=max_row
                        )

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

wb.active.add_chart(barchart, 'B6')
barchart.title = 'Payment Used Channel'
barchart.style = 2

wb.save(output_file)

def send_report_to_discord(webhookUrl, output):
    import discord
    from discord import SyncWebhook

    webhook = SyncWebhook.from_url(webhookUrl)

    with open(file=output, mode='rb') as file:
        excel_file = discord.File(file)
    
    webhook.send('This is an automated report', username='IT OPS BOT', file=excel_file)


send_report_to_discord(webhook_url, output_file)
